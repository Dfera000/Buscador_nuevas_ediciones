import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
import shutil
import time
import re
import unidecode
import requests
import gradio as gr
import tempfile
from dateutil import parser as du
import urllib.parse
import traceback
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
import os

# --- Constantes y Globales ---
STOPWORDS = {"y", "de", "la", "el", "los", "las", "en", "del", "un", "una", "unos", "unas", "por", "para"}
PAUSE_OPENLIBRARY = 0.8

log_messages = [] # Se mantiene como un acumulador global
driver_cultura_global = None
cultura_cookies_accepted_global = False

def _init_cultura_driver_for_spaces():
    """
    Inicializa un Chrome/Chromium headless compatible con Hugging Face Spaces.
    Detecta binarios y rutas típicas de Debian/Ubuntu (chromium/chromedriver).
    """
    chrome_options = Options()
    # Headless moderno y flags recomendados para contenedores/CI
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--log-level=3")

    # Localiza binario de Chromium en el Space
    chromium_candidates = [
        "/usr/bin/chromium",
        "/usr/bin/chromium-browser",
        shutil.which("chromium"),
        shutil.which("chromium-browser"),
    ]
    for path in chromium_candidates:
        if path:
            chrome_options.binary_location = path
            break

    # Localiza ChromeDriver en el Space
    chromedriver_candidates = [
        "/usr/bin/chromedriver",
        "/usr/lib/chromium-browser/chromedriver",
        shutil.which("chromedriver"),
    ]
    chromedriver_path = next((p for p in chromedriver_candidates if p), None)

    if not chromedriver_path:
        raise RuntimeError("No se encontró chromedriver en el sistema.")

    service = Service(chromedriver_path)
    return webdriver.Chrome(service=service, options=chrome_options)


# --- Funciones de Logging ---
def log(message):
    print(message)
    log_messages.append(str(message))

# --- [EL RESTO DE FUNCIONES DE LIMPIEZA Y BÚSQUEDA PERMANECEN EXACTAMENTE IGUAL] ---
def clean_year_value(year_str):
    if pd.isnull(year_str) or not isinstance(year_str, str): return None
    match_bracket_paren = re.search(r'[\[\(](\d{4})[\]\)]', year_str)
    if match_bracket_paren:
        try:
            year_cand = int(match_bracket_paren.group(1))
            if 1700 <= year_cand <= 2100: return year_cand
        except ValueError: pass
    match = re.search(r'\b(1[7-9]\d{2}|20\d{2}|2100)\b', year_str)
    if match:
        try: return int(match.group(1))
        except ValueError: pass
    return None

def select_priority_isbn(isbn_field_str):
    if pd.isnull(isbn_field_str) or not isinstance(isbn_field_str, str) or not isbn_field_str.strip(): return ""
    potential_isbns = [s.strip() for s in re.split(r'[;\s,]+', isbn_field_str) if s.strip()]
    isbn13_starts_with_9_strict, isbn13_general, isbn10_valid = [], [], []
    for isbn in potential_isbns:
        cleaned_isbn_text = re.sub(r'^(ISBN|isbn)\s*:\s*', '', isbn, flags=re.IGNORECASE)
        cleaned_isbn_digits = cleaned_isbn_text.replace('-', '').replace(' ', '').strip()
        if cleaned_isbn_digits.isdigit():
            if len(cleaned_isbn_digits) == 13:
                if cleaned_isbn_digits.startswith('9'): isbn13_starts_with_9_strict.append(cleaned_isbn_digits)
                if cleaned_isbn_digits.startswith(('978', '979')): isbn13_general.append(cleaned_isbn_digits)
            elif len(cleaned_isbn_digits) == 10: isbn10_valid.append(cleaned_isbn_digits)
    if isbn13_starts_with_9_strict: return isbn13_starts_with_9_strict[0]
    if isbn13_general: return isbn13_general[0]
    if isbn10_valid: return isbn10_valid[0]
    return ""

def clean_title_and_author_general(row_series):
    try:
        title_input = row_series.get('Title')
        if pd.isnull(title_input) or str(title_input).strip() == "":
            title_clean = "No disponible"
        else:
            title = unidecode.unidecode(str(title_input))
            title = re.sub(r'[^a-zA-Z0-9\s]', '', title)
            title_words = [word for word in title.split() if word.lower() not in STOPWORDS]
            title_clean = ' '.join(title_words[:5])
            if not title_clean.strip():
                title_clean = "No disponible"

        author_input = row_series.get('Author')
        if pd.isnull(author_input) or str(author_input).strip() == "":
            author_clean = "No disponible"
        else:
            author = unidecode.unidecode(str(author_input))
            author = re.sub(r'[^a-zA-Z\s]', '', author)
            author_parts = author.split()
            author_clean = ' '.join(author_parts[:2])
            if not author_clean.strip():
                author_clean = "No disponible"

        return title_clean, author_clean
    except Exception as e:
        log(f"Error en clean_title_and_author_general: {str(e)}")
        return "Error Limpieza", "Error Limpieza"

def clean_title_for_cultura_gob_search(original_title):
    if pd.isnull(original_title) or str(original_title).strip() == "":
        return "No disponible"
    try:
        title = unidecode.unidecode(str(original_title).lower())
        title = re.sub(r'[^a-zA-Z0-9\s]', '', title)
        title_words = title.split()
        title_clean_cultura = ' '.join(title_words[:5])
        if not title_clean_cultura.strip():
            return "No disponible"
        return title_clean_cultura
    except Exception as e:
        log(f"Error en clean_title_for_cultura_gob_search: {str(e)}")
        return "Error Limpieza Titulo CG"

def search_book_cultura_gob(driver, title_for_search, author_for_search, cookies_already_accepted):
    global cultura_cookies_accepted_global
    log(f"Cultura.gob: Buscando T='{title_for_search}', A='{author_for_search}'")

    if not driver:
        log("Cultura.gob: Driver no disponible. Saltando búsqueda.")
        return "Driver Error", None, None, None, None
    try:
        current_url = driver.current_url
        base_search_url = 'https://www.cultura.gob.es/webISBN/tituloSimpleFilter.do'
        if base_search_url not in current_url :
             driver.get(base_search_url + '?cache=init&prev_layout=busquedaisbn&layout=busquedaisbn&language=es')

        wait = WebDriverWait(driver, 20)

        if not cookies_already_accepted:
            try:
                cookie_xpaths = ["//button[contains(translate(normalize-space(.), 'ACEPTAR', 'aceptar'), 'aceptar')]", "//button[contains(translate(text(), 'ACEPTAR', 'aceptar'), 'Aceptar')]", "//a[contains(translate(normalize-space(.), 'ACEPTAR', 'aceptar'), 'aceptar')]"]
                for xpath in cookie_xpaths:
                    try:
                        cookie_button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                        driver.execute_script("arguments[0].click();", cookie_button)
                        cultura_cookies_accepted_global = True; time.sleep(1.0)
                        break
                    except: continue
                if not cultura_cookies_accepted_global: cultura_cookies_accepted_global = True
            except Exception as e_cookie:
                log(f"Cultura.gob: Advertencia (o banner no presente) al manejar cookies: {str(e_cookie)}")
                if not cultura_cookies_accepted_global: cultura_cookies_accepted_global = True

        search_box_xpath = "//input[@id='params.liConceptosExt[0].texto']"
        search_box = wait.until(EC.presence_of_element_located((By.XPATH, search_box_xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", search_box)
        time.sleep(0.3); search_box.clear()

        search_query = " ".join([part for part in [title_for_search, author_for_search] if part and part.strip()]) or title_for_search
        if not search_query.strip(): return "Query Vacía", None, None, None, None

        search_box.send_keys(search_query)
        submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and @value='Buscar']")))
        driver.execute_script("arguments[0].click();", submit_button)

        resultados_xpath, no_resultados_xpath = "//div[@class='isbnResultado']", "//div[@id='aviso']"
        try:
            WebDriverWait(driver, 12).until(EC.any_of(EC.presence_of_element_located((By.XPATH, resultados_xpath)), EC.presence_of_element_located((By.XPATH, no_resultados_xpath))))
        except TimeoutException:
            log("Cultura.gob: Timeout esperando resultados.")
            return "Timeout Resultados", None, None, None, None

        resultados_elements = driver.find_elements(By.XPATH, resultados_xpath)
        if not resultados_elements: return "No hallado", None, None, None, None

        libros = []
        for res_element in resultados_elements:
            titulo_text, autor_text, isbn_text, ano_val = "No disponible", "No disponible", "No disponible", None
            try:
                titulo_el = res_element.find_element(By.XPATH, ".//div[@class='isbnResDescripcion']//a[contains(@href, 'tituloDetalle')]")
                titulo_text = titulo_el.text.strip()
                try:
                    autor_p = res_element.find_element(By.XPATH, ".//div[@class='isbnResDescripcion']//p[contains(normalize-space(.), 'Autor/es:')]")
                    autor_match = re.search(r'Autor/es:\s*(.*?)(?:\n|$|F\. Edición:|ISBN:)', autor_p.text, re.DOTALL)
                    if autor_match: autor_text = re.sub(r'\s+', ' ', autor_match.group(1).strip())
                except: pass
                try:
                    isbn_el = res_element.find_element(By.XPATH, ".//div[@class='camposCheck']//a[contains(@href, 'tituloDetalle')] | .//div[contains(@class, 'camposIsbnRes')]//span[@class='isbn'][1]//strong")
                    isbn_text = isbn_el.text.strip()
                except: pass

                full_desc_text = " ".join([p.text for p in res_element.find_elements(By.XPATH, ".//div[@class='isbnResDescripcion']//p")])
                year_patterns = [r'\((\d{4})\)', r'F\.\s*Edición:\s*\D*(\d{4})\b', r'F\.\s*Publicación:\s*\D*(\d{4})\b', r'\b(1[89]\d{2}|20\d{2})\b']
                for pattern in year_patterns:
                    match = re.search(pattern, full_desc_text, re.IGNORECASE)
                    if match:
                        potential_year = int(match.group(1))
                        if 1800 <= potential_year <= 2100: ano_val = potential_year; break

                if ano_val: libros.append({'titulo': titulo_text, 'autor': autor_text, 'isbn': isbn_text, 'ano_edicion': ano_val})
            except Exception as e_proc_res:
                log(f"Cultura.gob: Error procesando resultado individual: {str(e_proc_res)}")

        if libros:
            libros.sort(key=lambda x: x.get('ano_edicion', 0), reverse=True)
            lr = libros[0]
            return "OK", lr['titulo'], lr['autor'], lr['isbn'], str(lr['ano_edicion'])
        return "No hallado (s/año)", None, None, None, None
    except Exception as e_sel:
        log(f"Cultura.gob: Error inesperado en Selenium: {str(e_sel)}\n{traceback.format_exc()}")
        return "Error Inesperado", None, None, None, None

def y_ol(date_input):
    if not date_input: return None
    date_str = str(date_input[0]) if isinstance(date_input, list) and date_input else str(date_input)
    try:
        if re.fullmatch(r'\d{4}', date_str):
            year = int(date_str)
            if 1700 <= year <= 2100: return year
    except: pass
    try: return du.parse(date_str, fuzzy=True, ignoretz=True).year
    except:
        m = re.search(r"\b(1[7-9]\d{2}|20\d{2}|2100)\b", date_str)
        return int(m.group(1)) if m else None

def g_ol(url, **kv):
    try:
        r = requests.get(url, timeout=20, **kv)
        r.raise_for_status(); return r
    except requests.exceptions.RequestException as e_req: log(f"OL Error GET: {e_req}"); return None
    except Exception as e_gen: log(f"OL Error general g_ol: {e_gen}"); return None

def works_from_isbn_ol(isbn):
    r = g_ol(f"https://openlibrary.org/isbn/{isbn}.json")
    return [w["key"] for w in r.json().get("works", [])] if r and r.content else []

def authors_of_work_ol(wk):
    r = g_ol(f"https://openlibrary.org{wk}.json"); names = []
    if r and r.content:
        for a_data in r.json().get("authors", []):
            author_entry = a_data.get("author", {}) if "author" in a_data else a_data
            author_key_path = author_entry.get("key")
            if author_key_path:
                rk_author = g_ol("https://openlibrary.org" + author_key_path + ".json")
                if rk_author and rk_author.content: names.append(rk_author.json().get("name", ""))
    return names

def eds_of_work_ol(wk, names_of_work_authors):
    url_base, params = f"https://openlibrary.org{wk}/editions.json", {"limit": 50, "fields": "key,title,publish_date,publish_year,isbn_13,isbn_10,identifiers,authors,author_name"}
    all_eds, current_offset, max_eds = [], 0, 150
    while len(all_eds) < max_eds:
        params["offset"] = current_offset
        r = g_ol(url_base, params=params);
        if not r or not r.content: break
        try: data = r.json()
        except: break
        current_entries = data.get("entries", [])
        if not current_entries: break
        for entry in current_entries:
            final_authors = entry.get("author_name", []) or names_of_work_authors
            all_eds.append({**entry, "author_list_resolved": final_authors})
        current_offset += len(current_entries)
        if len(current_entries) < params["limit"]: break
        time.sleep(PAUSE_OPENLIBRARY / 4)
    return all_eds

def search_editions_ol(title_for_query, author_for_query_hint=""):
    if not title_for_query or title_for_query == "No disponible": return []
    cleaned_title = title_for_query.replace('"', ''); q_parts = [f'title:"{cleaned_title}"']
    if author_for_query_hint and author_for_query_hint != "No disponible":
        cleaned_author = author_for_query_hint.replace('"', ''); q_parts.append(f'author:"{cleaned_author}"')
    q, fields = " AND ".join(q_parts), "key,title,author_name,publish_year,publish_date,isbn,first_publish_year"
    url = f"https://openlibrary.org/search.json?q={urllib.parse.quote_plus(q)}&fields={fields}&limit=10"
    r = g_ol(url); out = []
    if r and r.content:
        try:
            for d in r.json().get("docs", []):
                pub_year = None
                p_years_data = d.get("publish_year")
                if p_years_data:
                    valid_p_years = [y for y in p_years_data if isinstance(y, (int,float)) or (isinstance(y,str) and y.isdigit())]
                    if valid_p_years: pub_year = max(int(y) for y in valid_p_years)
                p_dates_data = d.get("publish_date")
                if not pub_year and p_dates_data:
                    p_dates_list = p_dates_data if isinstance(p_dates_data, list) else [p_dates_data]
                    parsed_years_from_dates = [y_ol(ds) for ds in p_dates_list if y_ol(ds)]
                    if parsed_years_from_dates: pub_year = max(parsed_years_from_dates)
                fp_year_data = d.get("first_publish_year")
                if not pub_year and fp_year_data and str(fp_year_data).isdigit(): pub_year = int(fp_year_data)
                out.append({"key": d.get("key"), "title": d.get("title"), "publish_date": str(pub_year) if pub_year else "", "author_list_resolved": d.get("author_name",[]), "isbn_candidate": d.get("isbn",[])})
        except Exception as e_parse_ol_search:
            log(f"OL: Error parseando resultados de búsqueda: {e_parse_ol_search}")
    return out

def author_ok_ol(targets, ol_authors):
    if not targets or (len(targets)==1 and targets[0]=="No disponible"): return True
    if not ol_authors: return False
    targets_p = [unidecode.unidecode(str(t)).lower().strip() for t in targets if t and str(t).strip()]
    ol_authors_p = [unidecode.unidecode(str(a)).lower().strip() for a in ol_authors if a and str(a).strip()]
    if not targets_p: return True
    for t_full in targets_p:
        t_parts = [p for p in t_full.split() if p]
        if not t_parts: continue
        for ol_full in ol_authors_p:
            if not ol_full: continue
            if t_full == ol_full or all(p in ol_full for p in t_parts): return True
            if t_parts[-1] in ol_full.split() and (len(t_parts)==1 or any(np in ol_full for np in t_parts[:-1])): return True
    return False

def best_edition_ol(original_isbn_cleaned, title_clean_general, author_clean_general):
    log(f"OL: Buscando T='{title_clean_general}', A='{author_clean_general}'")
    all_eds, work_authors = [], set()
    time.sleep(PAUSE_OPENLIBRARY / 3)

    if original_isbn_cleaned and original_isbn_cleaned != "No disponible":
        work_keys = works_from_isbn_ol(original_isbn_cleaned)
        if work_keys:
            for wk_idx, wk in enumerate(work_keys[:1]):
                if wk_idx > 0: time.sleep(PAUSE_OPENLIBRARY / 4)
                authors_wk = authors_of_work_ol(wk)
                if authors_wk: work_authors.update(authors_wk)
                eds_wk = eds_of_work_ol(wk, authors_wk or list(work_authors))
                if eds_wk: all_eds.extend(eds_wk)

    if title_clean_general and title_clean_general != "No disponible":
        time.sleep(PAUSE_OPENLIBRARY / 3)
        author_hint = author_clean_general if author_clean_general != "No disponible" else ""
        eds_title = search_editions_ol(title_clean_general, author_hint)
        if eds_title: all_eds.extend(eds_title)

    if not all_eds:
        return "No hallado", None, None, None, None

    filter_authors = [author_clean_general] if author_clean_general != "No disponible" else list(work_authors)
    best = {"year": -1, "title": None, "isbn": None, "author": None}

    for e_data in all_eds:
        if filter_authors and not author_ok_ol(filter_authors, e_data.get("author_list_resolved", [])): continue
        year = y_ol(e_data.get("publish_date") or e_data.get("publish_year"))
        if not year or year < 1700: continue
        if year > best["year"]:
            best["year"] = year; best["title"] = e_data.get("title"); chosen_isbn = None
            isbns13_data = e_data.get("isbn_13", []); isbns10_data = e_data.get("isbn_10", []); isbns_cand_data = e_data.get("isbn_candidate", [])
            if isbns13_data and isinstance(isbns13_data, list) and isbns13_data: chosen_isbn = isbns13_data[0]
            if not chosen_isbn and isbns10_data and isinstance(isbns10_data, list) and isbns10_data: chosen_isbn = isbns10_data[0]
            if not chosen_isbn and isbns_cand_data and isinstance(isbns_cand_data, list) and isbns_cand_data:
                for ic in isbns_cand_data:
                    if len(str(ic).replace("-","")) == 13: chosen_isbn = ic; break
                if not chosen_isbn and isbns_cand_data: chosen_isbn = isbns_cand_data[0]
            best["isbn"] = str(chosen_isbn).replace("-","") if chosen_isbn else None
            best["author"] = ", ".join(e_data.get("author_list_resolved", []) or filter_authors or ["Desconocido"])

    if best["year"] > -1:
        return "OK", best['title'], best['author'], best['isbn'], str(best['year'])
    else:
        return "No hallado (s/criterio)", None, None, None, None

# --- Función Principal de Procesamiento (MODIFICADA A GENERADOR) ---
def process_excel_generator(file_path_or_obj):
    global driver_cultura_global, cultura_cookies_accepted_global

    try:
        log("=======================================\nInicio del procesamiento del archivo Excel.\n=======================================")
        yield "\n".join(log_messages)

        try:
            df_peek = pd.read_excel(file_path_or_obj, nrows=None)
        except Exception as e_peek:
            log(f"Error Crítico al intentar leer el Excel: {e_peek}")
            yield "\n".join(log_messages)
            return

        # --- Inicialización del Driver (si es necesario) ---
        if 'Idioma' in df_peek.columns and 'es' in df_peek['Idioma'].astype(str).str.lower().unique():
            if not driver_cultura_global:
                log("Inicializando driver de Selenium para Cultura.gob...")
                yield "\n".join(log_messages)
                try:
                    chrome_options = Options(); chrome_options.add_argument("--headless"); chrome_options.add_argument("--no-sandbox"); chrome_options.add_argument("--disable-dev-shm-usage"); chrome_options.add_argument("--window-size=1920,1080"); chrome_options.add_argument("--disable-gpu"); chrome_options.add_argument("--log-level=3")
                    driver_cultura_global = _init_cultura_driver_for_spaces()
                    cultura_cookies_accepted_global = False
                    log("Driver de Cultura.gob inicializado.")
                    yield "\n".join(log_messages)
                except Exception as e_driver_init:
                    log(f"CRITICAL: No se pudo inicializar el driver de Cultura.gob: {e_driver_init}")
                    driver_cultura_global = None
                    yield "\n".join(log_messages)
        else:
            log("No hay libros en 'es' o 'Idioma' no presente, no se inicializa driver para Cultura.gob.")
            yield "\n".join(log_messages)

        log("Cargando archivo Excel...")
        df = pd.read_excel(file_path_or_obj)
        log(f"Archivo cargado. {len(df)} filas. Preparando búsqueda, espere unos segundos...")
        yield "\n".join(log_messages)
        original_excel_cols = list(df.columns)
        # Extraer el nombre del archivo original sin la ruta y la extensión
        base_name = os.path.basename(file_path_or_obj)
        name_without_ext = os.path.splitext(base_name)[0]
        # Crear el nuevo nombre para el archivo de salida
        output_file_name = f"Resultados_{name_without_ext}.xlsx"

        # --- Preparación del DataFrame ---
        if 'year' in df.columns: df['Year_cleaned_from_input'] = df['year'].apply(lambda x: clean_year_value(str(x) if pd.notnull(x) else None))
        else: df['Year_cleaned_from_input'] = pd.NA
        for col in ['Title', 'Author', 'Idioma', 'ISBN']:
            if col not in df.columns: df[col] = pd.NA
        df['ISBN_prioritario_input'] = df['ISBN'].apply(lambda x: select_priority_isbn(str(x) if pd.notnull(x) else None))
        res_cols = ['Título encontrado', 'Autor encontrado', 'ISBN encontrado', 'Año de edición encontrado']
        search_terms_cols = ['Título usado para búsqueda', 'Autor usado para búsqueda']
        for col in res_cols + search_terms_cols + ['Resultado']: df[col] = ""

# --- Bucle Principal de Procesamiento ---
        for index, row in df.iterrows():
            log(f"--- Fila Excel {index+1}/{len(df)} ---")
            original_title_excel, raw_author_excel = str(row.get('Title', '')), row.get('Author')
            idioma_excel, isbn_prioritario = str(row.get('Idioma', '')).strip().lower(), str(row.get('ISBN_prioritario_input', '')).strip()
            year_input_cleaned = pd.to_numeric(row.get('Year_cleaned_from_input'), errors='coerce')
            if pd.isna(year_input_cleaned):
                log("  Advertencia: Falta el año en la columna 'year'. Saltando fila.")
                log("") # Espacio en blanco
                df.at[index, 'Resultado'] = "Fallo - Input: Falta el año"
                yield "\n".join(log_messages)
                continue # Salta al siguiente libro del bucle
            author_post_comma = str(raw_author_excel).split(',', 1)[0].strip() if pd.notna(raw_author_excel) and ',' in str(raw_author_excel) else (str(raw_author_excel).strip() if pd.notna(raw_author_excel) else None)

            status, res_t, res_a, res_i, res_y, final_result_message = "No procesado", "", "", "", "", "No procesado"
            titulo_usado_para_busqueda_display, autor_usado_para_busqueda_display = "N/A", "N/A"

            if not original_title_excel.strip():
                final_result_message = "Fallo - Input: Título vacío"
            elif idioma_excel == 'es':
                titulo_busqueda_cultura = clean_title_for_cultura_gob_search(original_title_excel)
                _, autor_busqueda_cultura_temp = clean_title_and_author_general(pd.Series({'Title': '', 'Author': author_post_comma}))
                autor_busqueda_cultura = "" if autor_busqueda_cultura_temp in ["No disponible", "Error Limpieza"] else autor_busqueda_cultura_temp
                titulo_usado_para_busqueda_display, autor_usado_para_busqueda_display = titulo_busqueda_cultura, autor_busqueda_cultura or "N/A"
                if "No disponible" in titulo_busqueda_cultura:
                    final_result_message = "Fallo - Input: Título inválido"
                else: 
                    status, res_t, res_a, res_i, res_y = search_book_cultura_gob(driver_cultura_global, titulo_busqueda_cultura, autor_busqueda_cultura, cultura_cookies_accepted_global)
                    
                    # Si la búsqueda en Cultura.gob falla, intentamos con Open Library como respaldo
                    if status != "OK":
                        log(f"  -> Fallo en Cultura.gob ({status}). Intentando búsqueda de respaldo en Open Library...")
                        yield "\n".join(log_messages)
                        
                        titulo_busqueda_ol, autor_busqueda_ol_temp = clean_title_and_author_general(pd.Series({'Title': original_title_excel, 'Author': author_post_comma}))
                        autor_busqueda_ol = "" if autor_busqueda_ol_temp in ["No disponible", "Error Limpieza"] else autor_busqueda_ol_temp
                        
                        titulo_usado_para_busqueda_display = titulo_busqueda_ol
                        autor_usado_para_busqueda_display = autor_busqueda_ol or "N/A"

                        status, res_t, res_a, res_i, res_y = best_edition_ol(isbn_prioritario, titulo_busqueda_ol, autor_busqueda_ol)
                        time.sleep(PAUSE_OPENLIBRARY)

                        if status == "OK":
                            status = "OK_FALLBACK"
            elif idioma_excel == 'no-es':
                titulo_busqueda_ol, autor_busqueda_ol_temp = clean_title_and_author_general(pd.Series({'Title': original_title_excel, 'Author': author_post_comma}))
                autor_busqueda_ol = "" if autor_busqueda_ol_temp in ["No disponible", "Error Limpieza"] else autor_busqueda_ol_temp
                titulo_usado_para_busqueda_display, autor_usado_para_busqueda_display = titulo_busqueda_ol, autor_busqueda_ol or "N/A"
                if "No disponible" in titulo_busqueda_ol:
                    final_result_message = "Fallo - Input: Título inválido"
                else:
                    status, res_t, res_a, res_i, res_y = best_edition_ol(isbn_prioritario, titulo_busqueda_ol, autor_busqueda_ol)
                time.sleep(PAUSE_OPENLIBRARY)
            else:
                final_result_message = "Fallo - Input: Idioma Inválido"

            # --- BLOQUE DE PROCESAMIENTO DE RESULTADO (CORREGIDO) ---
            if status.startswith("OK"): # Captura "OK" y "OK_FALLBACK"
                year_found = pd.to_numeric(res_y, errors='coerce')
                warnings_list = []
                title_differs = False
                
                # Lógica de comparación de títulos unificada
                clean_found_title, _ = clean_title_and_author_general(pd.Series({'Title': res_t, 'Author': ''}))
                clean_search_title, _ = clean_title_and_author_general(pd.Series({'Title': original_title_excel, 'Author': ''}))
                if clean_found_title != clean_search_title.lower():
                     title_differs = True

                if title_differs: warnings_list.append("Título difiere")
                if autor_usado_para_busqueda_display == "N/A": warnings_list.append("Sin autor")
                if pd.isna(year_input_cleaned): warnings_list.append("Sin año de comparación")
                
                is_newer = pd.notna(year_found) and pd.notna(year_input_cleaned) and year_found > year_input_cleaned
                base_message = "Éxito, ed. más actual" if is_newer else "Éxito - sin versión más reciente"
                
                # Añade una etiqueta si el éxito vino de la búsqueda de respaldo
                if status == "OK_FALLBACK":
                    base_message = f" {base_message}"

                final_result_message = f"{base_message} - {', '.join(warnings_list)}" if warnings_list else base_message
            elif final_result_message == "No procesado":
                final_result_message = f"Fallo - {status}"

            df.loc[index, ['Título usado para búsqueda', 'Autor usado para búsqueda', 'Título encontrado', 'Autor encontrado', 'ISBN encontrado', 'Año de edición encontrado', 'Resultado']] = \
            [titulo_usado_para_busqueda_display, autor_usado_para_busqueda_display, res_t, res_a, res_i, res_y, final_result_message]
            log(f"  Resultado fila {index+1}: {final_result_message}")
            log("") # Línea de separación
            yield "\n".join(log_messages)

        log("=======================================\nProcesamiento de filas completado.\n=======================================")
        yield "\n".join(log_messages)

        # --- Creación y formato del archivo Excel de salida ---
        final_output_columns = [c for c in original_excel_cols if c in df.columns]
        extra_cols = ['Year_cleaned_from_input', 'ISBN_prioritario_input', 'Título usado para búsqueda', 'Autor usado para búsqueda', 'Título encontrado', 'Autor encontrado', 'ISBN encontrado', 'Año de edición encontrado', 'Resultado']
        final_output_columns.extend([c for c in extra_cols if c not in final_output_columns])
        df_output_final = df[final_output_columns]

        # -- Bloque de guardado y formato con la indentación corregida --

        # Guardar el DataFrame inicial en el archivo con el nuevo nombre
        df_output_final.to_excel(output_file_name, index=False, engine='openpyxl')

        # Cargar el archivo recién guardado para aplicarle estilos
        workbook = openpyxl.load_workbook(output_file_name)
        worksheet = workbook.active

        # Definición de estilos
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        # Localizar la columna de Resultado
        header = {cell.value: i+1 for i, cell in enumerate(worksheet[1])}
        result_col_idx = header.get('Resultado')

        # Aplicar estilos fila por fila
        if result_col_idx:
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=result_col_idx)
                resultado_str = cell.value or ""
                fill_to_apply = None

                if resultado_str.startswith("Fallo"):
                    fill_to_apply = red_fill
                elif "sin versión más reciente" in resultado_str:
                    fill_to_apply = None
                elif resultado_str == "Éxito, ed. más actual":
                    fill_to_apply = green_fill
                elif resultado_str.startswith("Éxito, ed. más actual -"):
                    fill_to_apply = yellow_fill

                if fill_to_apply:
                    for c in worksheet[row_idx]:
                        c.fill = fill_to_apply

            # Aplicar borde a toda la columna de Resultado
            for row in worksheet.iter_rows(min_col=result_col_idx, max_col=result_col_idx, min_row=1):
                for cell in row:
                    cell.border = thick_border

        # Guardar los cambios de estilo en el mismo archivo
        workbook.save(output_file_name)

        log(f"Archivo Excel generado con los resultados: {output_file_name}")
        yield (output_file_name, "\n".join(log_messages))

    except Exception as e:
        log(f"Error CRÍTICO general: {str(e)}\n{traceback.format_exc()}")
        yield (None, "\n".join(log_messages))

# --- Interfaz Gradio ---
with gr.Blocks(theme=gr.themes.Soft()) as demo:
    gr.Markdown("# Buscador de Últimas Ediciones de Libros")
    gr.Markdown(
        "**Pasos:**\n"
        "1.  Sube tu archivo Excel con las columnas requeridas.\n"
        "2.  Presiona el botón 'Procesar Archivo Excel'. El programa buscará los libros españoles en la base de datos del ISBN (Ministerio de Cultura) y los extranjeros en la API de Open Library\n"
        "3.  Descarga el archivo generado con los resultados.\n\n"
        "**Columnas Requeridas:** `Title`, `year`, `Idioma` (`es` o `no-es`). **Muy recomendables:** `ISBN`, `Author`.\n\n"
        "**Colores de Salida en el Excel:**\n"
        "- **Verde**: Éxito total. Se encontró una edición más nueva sin problemas.\n"
        "- **Amarillo**: Éxito con problemas (ej. el título difiere, sin autor, etc.).\n"
        "- **Rojo**: Fallo. La búsqueda no se pudo completar o hubo un error de input.\n"
        "- **Sin Color**: Búsqueda correcta, pero la edición encontrada es igual o anterior a la proporcionada."
    )

    with gr.Row():
        excel_input = gr.File(label="Sube tu archivo Excel (.xlsx)", type="filepath", file_types=['.xlsx'])
        processed_file_output = gr.File(label="Descarga el archivo procesado")

    log_output = gr.Textbox(label="Log del Proceso", interactive=False, lines=15, max_lines=30, autoscroll=True)
    submit_button = gr.Button("Procesar Archivo Excel")

    def gradio_excel_processing_interface(gradio_file_object):
        global log_messages
        log_messages = []

        if gradio_file_object is None:
            yield None, "Por favor, sube un archivo Excel."
            return

        for update in process_excel_generator(gradio_file_object.name):
            if isinstance(update, tuple):
                file_path, log_text = update
                yield file_path, log_text
            else:
                log_text = update
                yield None, log_text

    submit_button.click(gradio_excel_processing_interface, inputs=excel_input, outputs=[processed_file_output, log_output])


if __name__ == '__main__':
    try:
        # En Spaces NO uses share=True (es para enlaces efímeros en local/Colab)
        demo.launch(debug=True)  # o incluso puedes omitir launch, ver README
    finally:
        if driver_cultura_global:
            log("Cerrando driver global de Cultura.gob al finalizar el script/demo.")
            try:
                driver_cultura_global.quit()
            except Exception as e_quit:
                log(f"Error al intentar cerrar el driver global: {e_quit}")
            driver_cultura_global = None